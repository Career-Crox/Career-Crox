
import os
from pathlib import Path
import openpyxl
import requests
from datetime import datetime, timedelta
import random
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parents[1]
load_dotenv(BASE_DIR / ".env")
SEED_FILE = Path(os.environ.get("SEED_XLSX_PATH", str(BASE_DIR / "sample_data" / "Career_Crox_Interns_Seed.xlsx")))
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

HEADERS = {
    "apikey": SUPABASE_SERVICE_ROLE_KEY,
    "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}


TABLE_COLUMNS = {
    "users": {"user_id","username","password","full_name","designation","role","recruiter_code","is_active","theme_name","updated_at"},
    "candidates": {"candidate_id","full_name","phone","qualification","location","experience","preferred_location","qualification_level","total_experience","relevant_experience","in_hand_salary","career_gap","documents_availability","communication_skill","relevant_experience_range","relevant_in_hand_range","submission_date","process","recruiter_code","recruiter_name","recruiter_designation","status","all_details_sent","interview_reschedule_date","is_duplicate","notes","resume_filename","recording_filename","created_at","updated_at"},
    "tasks": {"task_id","title","description","assigned_to_user_id","assigned_to_name","assigned_by_user_id","assigned_by_name","status","priority","due_date","created_at","updated_at"},
    "notifications": {"notification_id","user_id","title","message","category","status","metadata","created_at"},
    "jd_master": {"jd_id","job_title","company","location","experience","salary","notes","created_at"},
    "settings": {"setting_key","setting_value","notes","Instructions"},
    "notes": {"candidate_id","username","note_type","body","created_at"},
    "messages": {"sender_username","recipient_username","body","created_at"},
    "interviews": {"interview_id","candidate_id","jd_id","stage","scheduled_at","status","created_at"},
    "submissions": {"submission_id","candidate_id","jd_id","recruiter_code","status","submitted_at"},
}

def trim_to_columns(rows, table_name):
    allowed = TABLE_COLUMNS[table_name]
    out = []
    for row in rows:
        out.append({k: row.get(k, "") for k in allowed if k in row})
    return out

CLEAR_FILTERS = {
    "users": {"user_id": "not.is.null"},
    "candidates": {"candidate_id": "not.is.null"},
    "tasks": {"task_id": "not.is.null"},
    "notifications": {"notification_id": "not.is.null"},
    "jd_master": {"jd_id": "not.is.null"},
    "settings": {"setting_key": "not.is.null"},
    "interviews": {"interview_id": "not.is.null"},
    "submissions": {"submission_id": "not.is.null"},
    "notes": {"id": "gt.0"},
    "messages": {"id": "gt.0"},
}

def now_iso():
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

def clean_row(row):
    cleaned = {}
    for k, v in row.items():
        if k is None:
            continue
        if isinstance(v, datetime):
            cleaned[k] = v.isoformat()
        elif v is None:
            cleaned[k] = ""
        else:
            cleaned[k] = str(v).strip() if not isinstance(v, (int, float)) else str(v)
    return cleaned

def parse_sheet_rows(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value is not None else None for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        item = {}
        for idx, value in enumerate(row):
            key = headers[idx] if idx < len(headers) else None
            if not key:
                continue
            item[key] = value
        if item:
            rows.append(clean_row(item))
    wb.close()
    return rows

def derive_payload(xlsx_path):
    users = parse_sheet_rows(xlsx_path, "Users")
    candidates = parse_sheet_rows(xlsx_path, "Candidates")
    tasks = parse_sheet_rows(xlsx_path, "Tasks")
    notifications = parse_sheet_rows(xlsx_path, "Notifications")
    jds = parse_sheet_rows(xlsx_path, "JD_Master")
    settings = parse_sheet_rows(xlsx_path, "Settings")

    usernames = [u.get("username", "") for u in users]
    candidate_codes = [c.get("candidate_id", "") for c in candidates]
    now = datetime.now()

    notes = []
    for candidate_id, username, note_type, body, day_offset in [
        ("C001", "recruiter.01", "public", "Candidate confirmed interview slot.", -2),
        ("C003", "recruiter.02", "public", "Strong communication and ready for next round.", -1),
        ("C001", "recruiter.01", "private", "Responds better after 6 PM.", -1),
    ]:
        if candidate_id in candidate_codes and username in usernames:
            notes.append({
                "candidate_id": candidate_id,
                "username": username,
                "note_type": note_type,
                "body": body,
                "created_at": (now + timedelta(days=day_offset, hours=random.randint(8, 18))).isoformat(timespec="seconds")
            })

    messages = []
    for sender, recipient, body, day_offset in [
        ("admin", "tl.noida", "Please review pending Airtel profiles today.", -1),
        ("tl.noida", "recruiter.01", "Update note history after callback.", -1),
    ]:
        if sender in usernames and recipient in usernames:
            messages.append({
                "sender_username": sender,
                "recipient_username": recipient,
                "body": body,
                "created_at": (now + timedelta(days=day_offset, hours=random.randint(8, 18))).isoformat(timespec="seconds")
            })

    interviews = []
    submissions = []
    for idx, c in enumerate(candidates, start=1):
        process = c.get("process") or ""
        jd_match = next((j for j in jds if (j.get("company") or "").strip().lower() == process.strip().lower()), None)
        jd_id = jd_match.get("jd_id") if jd_match else ""
        submissions.append({
            "submission_id": f"S{idx:03d}",
            "candidate_id": c.get("candidate_id", f"C{idx:03d}"),
            "jd_id": jd_id,
            "recruiter_code": c.get("recruiter_code", ""),
            "status": c.get("status", "New"),
            "submitted_at": c.get("submission_date") or now.strftime("%Y-%m-%d")
        })
        if "interview" in (c.get("status") or "").lower() or c.get("interview_reschedule_date"):
            interviews.append({
                "interview_id": f"I{idx:03d}",
                "candidate_id": c.get("candidate_id", f"C{idx:03d}"),
                "jd_id": jd_id,
                "stage": c.get("status", "Screening"),
                "scheduled_at": c.get("interview_reschedule_date") or f"{c.get('submission_date') or now.strftime('%Y-%m-%d')} 11:00",
                "status": "Scheduled",
                "created_at": now_iso()
            })

    return {
        "users": trim_to_columns(users, "users"),
        "candidates": trim_to_columns(candidates, "candidates"),
        "tasks": trim_to_columns(tasks, "tasks"),
        "notifications": trim_to_columns(notifications, "notifications"),
        "jd_master": trim_to_columns(jds, "jd_master"),
        "settings": trim_to_columns(settings, "settings"),
        "notes": trim_to_columns(notes, "notes"),
        "messages": trim_to_columns(messages, "messages"),
        "interviews": trim_to_columns(interviews, "interviews"),
        "submissions": trim_to_columns(submissions, "submissions"),
    }

def req(method, path, params=None, body=None):
    resp = requests.request(method, f"{SUPABASE_URL}/rest/v1/{path}", headers=HEADERS, params=params or {}, json=body, timeout=30)
    if resp.status_code >= 400:
        raise RuntimeError(f"{method} {path} failed: {resp.status_code} {resp.text[:300]}")
    return resp.json() if resp.text else []

def main():
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise SystemExit("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY first.")
    if not SEED_FILE.exists():
        raise SystemExit(f"Seed file not found: {SEED_FILE}")

    payload = derive_payload(SEED_FILE)

    for table, params in CLEAR_FILTERS.items():
        req("DELETE", table, params=params)

    for table, rows in payload.items():
        if rows:
            req("POST", table, body=rows)

    print("Import complete.")

if __name__ == "__main__":
    main()
